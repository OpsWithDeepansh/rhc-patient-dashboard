from datetime import datetime
import json
import os
import re
import sys
from openpyxl import load_workbook
from tkinter import Tk
from tkinter.filedialog import askopenfilename

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")

MONTH_ORDER = {
    "Jan": 1,
    "Feb": 2,
    "Mar": 3,
    "Apr": 4,
    "May": 5,
    "Jun": 6,
    "Jul": 7,
    "Aug": 8,
    "Sep": 9,
    "Oct": 10,
    "Nov": 11,
    "Dec": 12
}


def sort_month_code(month_code):
    month_name = month_code[:3]
    year = int(month_code[3:])

    return year, MONTH_ORDER.get(month_name, 99)

# Hide Tkinter window
Tk().withdraw()

# Select Excel file
file_path = askopenfilename(
    title="Select Patient on Therapy Excel File",
    filetypes=[("Excel Files", "*.xlsx")]
)

if not file_path:
    print("No file selected.")
    sys.exit()

# Open workbook
wb = load_workbook(file_path, data_only=True)

# Open Helper sheet
if "Helper" not in wb.sheetnames:
    print("Could not find the Helper sheet in the selected workbook.")
    sys.exit()

ws = wb["Helper"]

# Find last agent row

last_agent_row = 2

while (
    ws[f"A{last_agent_row}"].value is not None
    and str(ws[f"A{last_agent_row}"].value).lower() != "total"
):
    last_agent_row += 1

last_agent_row -= 1

# Read KPI row (Row 8)

kpi = {
    "totalPatients": ws["B8"].value,
    "connected": ws["C8"].value,
    "notConnected": ws["D8"].value,
    "inactive": ws["E8"].value,
    "denied": ws["F8"].value,
    "pending": ws["G8"].value,
    "callableLeads": ws["H8"].value,
    "connectivity": round(ws["I8"].value * 100, 2)
}

agents = {}

for row in range(2, last_agent_row + 1):

    agent_name = ws[f"A{row}"].value

    agents[agent_name.lower()] = {
    "connected": ws[f"C{row}"].value,
    "notConnected": ws[f"D{row}"].value,
    "inactive": ws[f"E{row}"].value,
    "denied": ws[f"F{row}"].value,
    "pending": ws[f"G{row}"].value,
    "callableLeads": ws[f"H{row}"].value,
    "connectivity": round(
    ws[f"I{row}"].value * 100,
    2
)
    }

trend = {}

for row in range(2, last_agent_row + 1):

    agent_name = ws[f"A{row}"].value

    trend[agent_name.lower()] = round(
        ws[f"I{row}"].value * 100,
        2
    )

daily_performance = {
    "connected": [],
    "notConnected": []
}

for row in range(2, last_agent_row + 1):

    daily_performance["connected"].append(
        ws[f"M{row}"].value
    )

    daily_performance["notConnected"].append(
        ws[f"N{row}"].value
    )

itp = {
    "connected": [],
    "pending": []
}

for row in range(2, last_agent_row + 1):

    itp["connected"].append(
        ws[f"S{row}"].value
    )

    itp["pending"].append(
        ws[f"U{row}"].value
    )

dashboard_data = {
    "totalPatients": kpi["totalPatients"],
    "connected": kpi["connected"],
    "notConnected": kpi["notConnected"],
    "inactive": kpi["inactive"],
    "denied": kpi["denied"],
    "pending": kpi["pending"],
    "callableLeads": kpi["callableLeads"],
    "connectivity": kpi["connectivity"],

    "agents": agents,

    "trend": trend,

    "dailyPerformance": daily_performance,

    "itp": itp
}

dashboard_data["generatedOn"] = datetime.now().strftime(
    "%d-%b-%Y %I:%M %p"
)

# Get filename

excel_name = os.path.basename(file_path)

# Extract Jun'26

match = re.search(r"([A-Za-z]{3})'(\d{2})", excel_name)

if match:

    month_name = match.group(1)
    year = match.group(2)

    json_file_name = f"{month_name}{year}.json"

    os.makedirs(DATA_DIR, exist_ok=True)

    output_path = os.path.join(
        DATA_DIR,
        json_file_name
    )

    with open(output_path, "w") as f:

        json.dump(
            dashboard_data,
            f,
            indent=4,
            ensure_ascii=False
        )

    print(f"\nJSON Created: {output_path}")

    # Update months.json

    months_file = os.path.join(
        DATA_DIR,
        "months.json"
    )

    months = []

    if os.path.exists(months_file):

        with open(months_file, "r") as f:
            months = json.load(f)

    month_code = f"{month_name}{year}"

    if month_code not in months:
        months.append(month_code)

    months.sort(key=sort_month_code)

    with open(months_file, "w") as f:

        json.dump(
            months,
            f,
            indent=4
        )

    print("months.json Updated")

    os.system("publish_dashboard.bat")

else:

    print(
        "Could not determine month from filename."
    )
